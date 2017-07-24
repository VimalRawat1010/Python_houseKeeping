#!/usr/bin/python
import math
import random
import re
from openpyxl import load_workbook
import numpy as np
import csv
import pandas as pd



# Reading XLSX file


# Method 1
wb = load_workbook('Analytics_case.xlsx', read_only=True)
print wb.sheetnames

Allusers = wb.get_sheet_by_name('All users')
Redeemers = wb.get_sheet_by_name('Redeemers')


# Method 2 Using pandas
df = pd.read_excel('Analytics_case.xlsx',  sheetname='All users')
print(df.columns)
Week = df['Week'].values


# print the column names
print (Week[0:10])

# Using map function
store1 =[10,15,20,23,53]
store2 =[8,21,23,12,45]
cheapest =map(min,store1,store2)
cheapest


# creating a lambda function
my_function =lambda a,b,c:a+b+c
my_function(1,2,3) 
people =['Dr. Vimal Rawat', 'Dr. John johny', 'Dr. Noman hooter']



# import numpy as np

my_list =[1,2,3,4]
x = np.array(my_list)
m = np.array([[1,3,4],[2,3,5]])
m
m.shape
n = np.arange(0,30,2)
n = n.reshape(3,5)

o = np.linspace(0,10,9)
o
np.ones((10,10))

m


# Iterating over np array

x = np.random.randint(0,10,(4,3))

for row in x:
    print row
    
for i in range(len(x)):
    print x[i]
    
for i,row in enumerate(x):
    print "row", i, 'is', row 
 
 
 
# iterate over 2 arrays with zip function
 
 
 