import math
import random
import re
from openpyxl import load_workbook
import numpy as np
import csv
import pandas as pd



# Reading XLSX file


# Method 1
wb = load_workbook('Analytics_case.xlsx', read_only=True)
print wb.sheetnames

Allusers = wb.get_sheet_by_name('All users')
Redeemers = wb.get_sheet_by_name('Redeemers')


# Method 2 Using pandas
df = pd.read_excel('Analytics_case.xlsx',  sheetname='All users')
print(df.columns)
userID = df['User ID'].values


# print the column names
print (userID[0:10])



# print (type(userID)) #print df.columns
# get the values for a given column
# values = df['Arm_id'].values
# get a data frame with selected columns
# FORMAT = ['Arm_id', 'DSPName', 'Pincode']
# df_selected = df[FORMAT]
